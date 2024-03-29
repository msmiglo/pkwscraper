
import re
from zipfile import ZipFile

from lxml import html
from openpyxl import load_workbook
import xlrd

from pkwscraper.lib.dbdriver import DbDriver
from pkwscraper.lib.downloader import Downloader
from pkwscraper.lib.scraper.base_scraper import BaseScraper


ELECTION_TYPE = "sejm"
YEAR = 2015
RAW_DATA_DIRECTORY = "./pkwscraper/data/sejm/2015/raw/"
RESCRIBED_DATA_DIRECTORY = "./pkwscraper/data/sejm/2015/rescribed/"
PREPROCESSED_DATA_DIRECTORY = "./pkwscraper/data/sejm/2015/preprocessed/"


class Sejm2015Scraper(BaseScraper):
    def __init__(self, db=None):
        # create downloader object
        self.dl = Downloader(year=2015, directory=RAW_DATA_DIRECTORY)

        # open db for rescribing
        if db is None:
            db = DbDriver(RESCRIBED_DATA_DIRECTORY)
        if not isinstance(db, DbDriver):
            raise TypeError("Please pass an instance of `DbDriver` or `None`.")
        if db.read_only:
            raise RuntimeError(
                "Please pass `DbDriver` for writing or `None`.")
        self.db = db

        # for checking
        self.all_votes = 0

    def run_all(self):
        # TODO - check, if the data is already downloaded
        #        and omit unnecessary steps
        self._download_voivodships()
        self._download_okregi()
        self._download_committees()
        self._download_xls_candidates()
        self._download_html_candidates()
        self._download_mandates_winners_and_powiaty()
        self._download_gminy_and_obwody()
        self._download_voting_results()

        print("dumping DB tables...")
        self.db.dump_tables()
        print("DB closed.")
        print()

    def _download_voivodships(self):
        relative_path = "/index.html"
        html_content = self.dl.download(relative_path)
        html_tree = html.fromstring(html_content)

        xpath_voivodships = '/html/body//div[@id="home"]//' \
                            'div[@class="home_content home_mapa"]//svg//a'
        voivodship_elems = html_tree.xpath(xpath_voivodships)
        print()
        print(len(voivodship_elems))

        self.db.create_table("województwa")

        for html_elem in voivodship_elems:
            href_path = html_elem.attrib['xlink:href']
            code = re.findall('\d+', href_path)[0]
            name = html_elem.getchildren()[0].attrib["rel"]
            geo = html_elem.getchildren()[0].attrib["d"]

            self.db["województwa"].put({
                "code": code,
                "name": name,
                "geo": geo
            })

            print(name, end=", ")
        print()

    def _download_okregi(self):
        relative_path = "/349_Wyniki_Sejm/0/0.html"
        html_content = self.dl.download(relative_path)
        html_tree = html.fromstring(html_content)

        xpath_all_votes = '/html/body/div/div[4]/div[2]/div[3]' \
                          '/div[2]/div/div[1]/div/div[3]/div[2]/text()'
        self.all_votes = int(html_tree.xpath(xpath_all_votes)[0])
        print()
        print(f"All votes: {self.all_votes}")

        xpath_okregi_mapa = '/html/body//div[@id="wyniki1"]//' \
                            'div[@id="wyniki1_top_mapa"]//svg//a'
        okregi_hrefs = html_tree.xpath(xpath_okregi_mapa)

        xpath_okregi_table = '/html/body//div[@id="wyniki1"]//' \
                             'div[@id="wyniki1_tabela_suma"]//table/tbody/tr'
        okregi_table_rows = html_tree.xpath(xpath_okregi_table)
        assert len(okregi_hrefs) == len(okregi_table_rows), \
            f"invalid number of constituencies: {len(okregi_hrefs)}, {len(okregi_table_rows)}"
        print()
        print(len(okregi_table_rows))

        self.db.create_table("okręgi")

        for html_map_elem, html_table_row_elem in \
                zip(okregi_hrefs, okregi_table_rows):
            href_path = html_map_elem.attrib['xlink:href']
            path_elem = html_map_elem.getchildren()[0]
            name = path_elem.attrib["rel"]
            number = re.findall('\d+', name)[0]
            geo = path_elem.attrib["d"]

            cell_elems = html_table_row_elem.getchildren()
            number_elem = cell_elems[0].getchildren()[0]
            href_path_2 = number_elem.attrib['href']
            number_2 = number_elem.text

            voivodship = cell_elems[1].getchildren()[0].text
            headquarters = cell_elems[2].getchildren()[0].text
            mandates = cell_elems[3].getchildren()[0].text

            assert href_path == href_path_2, f"invalid href: {href_path}, {href_path_2}"
            assert number == number_2, f"invalid number: {number}, {number_2}"

            self.db["okręgi"].put({
                "number": number,
                "headquarters": headquarters,
                "voivodship": voivodship,
                "mandates": mandates,
                "geo": geo
            })

            print(name, end=", ")
        print()

    def _download_committees(self):
        relative_path = "/komitety.html"
        html_content = self.dl.download(relative_path)
        html_tree = html.fromstring(html_content)

        xpath_committees = '/html/body//div[@id="komitety"]//' \
                            'table/tbody/tr'
        committees_elems = html_tree.xpath(xpath_committees)
        print()
        print(len(committees_elems))

        self.db.create_table("komitety")

        for html_elem in committees_elems:
            value_elems = [child.getchildren()[0]
                              for child in html_elem.getchildren()]
            href_path = value_elems[0].attrib['href']

            number = value_elems[0].text
            signature = value_elems[1].text
            type_ = value_elems[2].text
            name = value_elems[3].text
            shortname = value_elems[4].text
            sejm_candidates = value_elems[5].text
            senat_candidates = value_elems[6].text
            status = value_elems[7].text

            self.db["komitety"].put({
                "number": number,
                "signature": signature,
                "type": type_,
                "name": name,
                "shortname": shortname,
                "sejm_candidates": sejm_candidates,
                "senat_candidates": senat_candidates,
                "status": status
            })

            print(name, end=", ")
        print()

    def _download_xls_candidates(self):
        # download xls data
        self.dl.download("/kandydaci.zip")
        with ZipFile(RAW_DATA_DIRECTORY + "/kandydaci.zip") as zf:
            zf.extractall(RAW_DATA_DIRECTORY)

        # open xls
        book = xlrd.open_workbook(
            RAW_DATA_DIRECTORY + "/kandsejm2015-10-19-10-00.xls")
        sheet = book.sheet_by_index(0)

        self.db.create_table("kandydaci_xls")

        # iterate over candidates
        for row_index in range(1, sheet.nrows):
            row = sheet.row(row_index)

            okreg_number = row[0].value
            list_number = row[1].value
            committee_name = row[2].value
            position = row[3].value
            surname = row[4].value
            names = row[5].value
            gender = row[6].value
            residence = row[7].value
            occupation = row[8].value
            party = row[9].value

            self.db["kandydaci_xls"].put({
                "okreg_number": int(okreg_number),
                "list_number": int(list_number),
                "committee_name": committee_name,
                "position": int(position),
                "surname": surname,
                "names": names,
                "gender": gender,
                "residence": residence,
                "occupation": occupation,
                "party": party,
            })

        n_candidates = sheet.nrows - 1
        n_candidates_2 = len(self.db['kandydaci_xls'].find({}))
        assert n_candidates == n_candidates_2, \
               f"invalid candidates number: {n_candidates}, {n_candidates_2}"

        print()
        print(f"Found {n_candidates} candidates.")
        print()

    def _download_html_candidates(self):
        # enumerate constituencies
        relative_url_template = "/349_Wyniki_Sejm/0/0/{}.html"
        okregi = self.db["okręgi"].find({}, fields="number")
        self.db.create_table("kandydaci_html")

        # iterate over constituencies
        for okreg_number in okregi:
            # find candidates from constituency page
            relative_url = relative_url_template.format(okreg_number)
            html_content = self.dl.download(relative_url)
            html_tree = html.fromstring(html_content)

            xpath_candidates = '/html/body//div[@id="tresc"]//' \
                            'div[@id="wyniki1_tabela_frek"][2]//tbody/tr'
            candidates_elements = html_tree.xpath(xpath_candidates)

            # save records
            for row_elem in candidates_elements:
                cells = row_elem.getchildren()
                committee_name = cells[1].text_content()
                full_name = list(cells[3].itertext())[1]
                class_name = row_elem.get("class")
                crossed_out = (
                    class_name is not None
                    and "skreslony" in class_name
                )

                if crossed_out:
                    print(f"Crossed out candidate: {full_name},"
                          f" constituency: {okreg_number}, committee:"
                          f" {committee_name}.")

                self.db["kandydaci_html"].put({
                    "okreg_number": int(okreg_number),
                    "committee_shortname": committee_name,
                    "full_name": full_name,
                    "crossed_out": crossed_out
                })

    def _download_mandates_winners_and_powiaty(self):
        relative_url_template = "/349_Wyniki_Sejm/0/0/{}.html"
        okregi = self.db["okręgi"].find({}, fields="number")

        self.db.create_table("mandaty")
        self.db.create_table("powiaty")

        print()

        for okreg_number in okregi:
            relative_url = relative_url_template.format(okreg_number)
            html_content = self.dl.download(relative_url)
            html_tree = html.fromstring(html_content)

            xpath_powiaty = '/html/body//div[@id="tresc"]//' \
                            'div[@id="wyniki1_top_mapa"]//svg//a'
            powiaty = html_tree.xpath(xpath_powiaty)

            for powiat_elem in powiaty:
                href_path = powiat_elem.attrib['xlink:href']
                code = re.findall('\d+', href_path)[1]
                name = powiat_elem.getchildren()[0].attrib["rel"]
                geo = powiat_elem.getchildren()[0].attrib["d"]

                self.db["powiaty"].put({
                    "constituency_number": okreg_number,
                    "code": code,
                    "name": name,
                    "geo": geo
                })

            xpath_winners = '/html/body//div[@id="wyniki1_tabela_frek"][1]//' \
                            'table/tbody/tr'
            winners_elems = html_tree.xpath(xpath_winners)

            for row_elem in winners_elems:
                cells = row_elem.getchildren()
                constituency_number = list(cells[0].itertext())[1]
                list_number = cells[1].text_content()
                position = cells[2].text_content()
                committee_name = cells[3].text_content()
                full_name = list(cells[4].itertext())[1]
                assert okreg_number == constituency_number, \
                    f"invalid constituency number: {okreg_number}, {constituency_number}"

                self.db["mandaty"].put({
                    "constituency_number": constituency_number,
                    "list_number": list_number,
                    "position": position,
                    "committee_name": committee_name,
                    "full_name": full_name
                })

                print(full_name, end=", ")
        print()
        n_powiaty = len(self.db["powiaty"].find({}))
        n_mandates = len(self.db["mandaty"].find({}))
        print(f"Found {n_powiaty} districts.")
        print(f"{n_mandates} of mandates were given.")

    def _download_gminy_and_obwody(self):
        # get html content for all districts (powiaty)
        powiaty_codes = self.db["powiaty"].find({}, fields="code")
        relative_url_template = "/349_Wyniki_Sejm/0/1/0/{}.html"
        xpath_gminy = '/html/body//div[@id="tresc"]//' \
                      'div[@id="wyniki1_top_mapa"]//svg//a'

        self.db.create_table("gminy")
        self.db.create_table("obwody")

        # extract communes data for each district
        for powiat_code in powiaty_codes:
            relative_url = relative_url_template.format(powiat_code)
            html_content = self.dl.download(relative_url)
            html_tree = html.fromstring(html_content)

            gminy = html_tree.xpath(xpath_gminy)

            for gmina_elem in gminy:
                # get values of attributes
                href_path = gmina_elem.attrib['xlink:href']
                code = re.findall('\d+', href_path)[0]
                partial_name = gmina_elem.getchildren()[0].attrib["rel"]
                geo = gmina_elem.getchildren()[0].attrib["d"]

                # Łódź and Cracow cities are divided into city districts
                # but they have no polling districts assigned to them...
                # Skip them.
                # TODO - MAYBE IT WOULD BE BETTER TO GET RID OF THESE
                # RECORDS IN THE PREPROCESSING STEP...
                skip_condition = (
                    partial_name.startswith("Łódź")
                    or partial_name.startswith("Kraków")
                )
                skip_condition = skip_condition and code[-2:] != "01"
                if skip_condition:
                    continue

                # put record in DB
                self.db["gminy"].put({
                    "code": code,
                    "partial_name": partial_name,
                    "geo": geo
                })
                # NOTE - `name`, `rural_or_urban` will be taken from
                #        polling districts data in preprocessing step

        # extract polling districts information from xlsx
        self.dl.download("/wyniki_zb/2015-gl-lis-obw.zip")
        with ZipFile(
                RAW_DATA_DIRECTORY + "wyniki_zb_2015-gl-lis-obw.zip") as zf:
            zf.extractall(RAW_DATA_DIRECTORY)

        all_votes = 0
        book = xlrd.open_workbook(
            RAW_DATA_DIRECTORY + "/2015-gl-lis-obw.xls")
        sheet = book.sheet_by_index(0)
        for row_index in range(1, sheet.nrows):
            row = sheet.row(row_index)
            all_votes += int(row[27].value)

            self.db["obwody"].put({
                "constituency_number":          int(row[0].value),
                "senate_constituency_number":   int(row[1].value),
                "commune_code":                 row[2].value,
                "commune_name":                 row[3].value,
                "polling_district_number":      int(row[4].value),
                "full_address":                 row[5].value,
                "voters":                       int(row[6].value),
                "got_ballots":                  int(row[7].value),
                "unused_ballots":               int(row[8].value),
                "given_ballots":                int(row[9].value),
                "proxy_voters":                 int(row[10].value),
                "certificate_voters":           int(row[11].value),
                "voting_packets":               int(row[12].value),
                "return_envelopes":             int(row[13].value),
                "envelopes_without_statement":  int(row[14].value),
                "unsigned_statement":           int(row[15].value),
                "without_voting_envelope":      int(row[16].value),
                "unseeled_voting_envelopes":    int(row[17].value),
                "envelopes_accepted":           int(row[18].value),
                "ballots_from_box":             int(row[19].value),
                "envelopes_from_ballot_box":    int(row[20].value),
                "ballots_invalid":              int(row[21].value),
                "ballots_valid":                int(row[22].value),
                "votes_invalid":                int(row[23].value),
                "invalid_2_candidates":         int(row[24].value),
                "invalid_no_vote":              int(row[25].value),
                "invalid_candidate":            int(row[26].value),
                "votes_valid":                  int(row[27].value)
            })

        if not all_votes == self.all_votes:
            raise RuntimeError(
                "Incompatible votes sum: {all_votes} vs {self.all_votes}.")

        print()
        n_gminy = len(self.db["gminy"].find({}))
        print(f"Found {n_gminy} communes.")
        print()
        n_obwody = len(self.db["obwody"].find({}))
        print(f"Found {n_obwody} voting constituencies.")

    def _download_voting_results(self):
        """ Read votes from 41 xlsx files """
        print()
        # create extension of polling districts data
        self.db.create_table("obwody_uzupełnienie")
        all_votes = 0

        # iterate over constituencies
        table_name_template = "wyniki_{}"
        constituencies = self.db["okręgi"].find({}, fields="number")
        for constituency_number in constituencies:
            # create table with results for a constituency
            constituency_number = int(constituency_number)
            table_name = table_name_template.format(constituency_number)
            self.db.create_table(table_name)

            # read data
            relative_url = f"/wyniki_okr_sejm/{constituency_number:02d}.xlsx"
            self.dl.download(relative_url)
            filename = relative_url.replace("/", "_").strip("_")

            book = load_workbook(RAW_DATA_DIRECTORY + "/" + filename, read_only=True)
            sheet = book.active

            # read the first row with lists and candidates names
            last_cell_of_protocole_columns = sheet.cell(1, 27).value
            assert last_cell_of_protocole_columns == "Sejm - Liczba głosów ważnych oddanych łącznie na wszystkie listy kandydatów", \
                f"wrong columns alignment: {last_protocole_column_value}"
            current_list = None
            candidates = []
            # iterate over cells
            for column_index in range(27, sheet.max_column):
                value = sheet.cell(1, column_index+1).value
                # beginning of new list
                if not current_list:
                    current_list = value
                    candidates.append((current_list, "sum"))
                    continue
                # end of current list
                if value == f"Razem {current_list}":
                    candidates.append((current_list, "sum"))
                    current_list = None
                    continue
                # get candidate
                candidate_name = value
                candidates.append((current_list, candidate_name))

            # check correctness
            if current_list is not None:
                raise ValueError("Did not found end of list of candidates.")

            # iterate over polling districts (1 row - 1 polling district)
            print(f"Iterating over polling districts in constituency no. {constituency_number}...")
            for row_index, row in enumerate(sheet.iter_rows()):
                polling_district_votes = 0
                if row_index == 0:
                    continue
                if row_index % 100 == 0:
                    print(f"{row_index} of {sheet.max_row-1}",
                          end=", ", flush=True)

                # get additional polling district data
                commune_name = row[1].value
                commune_code = row[2].value
                commune_code = re.findall('\d+', commune_code)[0]
                commission_name = row[3].value
                polling_district_number = row[4].value

                # put it into table
                self.db["obwody_uzupełnienie"].put({
                    "commune_name": commune_name,
                    "commune_code": commune_code,
                    "commission_name": commission_name,
                    "polling_district_number": polling_district_number
                })

                # prepare list of cells corresponding with candidates
                candidates_cell_range = list(range(27, sheet.max_column))
                if len(candidates_cell_range) != len(candidates):
                    raise ValueError(
                        f"Different lenght of candidates lists: "
                        f"{len(candidates)}/{len(candidates_cell_range)}"
                    )

                # prepare polling district entry
                record = {
                    "commune_code": commune_code,
                    "polling_district_number": polling_district_number,
                    "candidates_count": 0,
                }

                # iterate over candidates in the given polling district
                for candidate, column_index in zip(
                        candidates, candidates_cell_range):
                    committee_name, candidate_name = candidate
                    votes = row[column_index].value
                    if candidate_name == "sum":
                        continue

                    try:
                        all_votes += int(votes)
                    except ValueError:
                        if votes != "XXXXX":
                            raise

                    # add field to the record
                    assert "/" not in committee_name + candidate_name, \
                           "`{committee_name + candidate_name}` has slash"
                    candidate_identifier = f"{committee_name}/{candidate_name}"
                    record[candidate_identifier] = votes
                    record["candidates_count"] += 1

                # add record to table
                self.db[table_name].put(record)

            print()
            print(f"Finished constituency no. {constituency_number}.")
        print()

        if not all_votes == self.all_votes:
            print(f"Incompatible votes sum: {all_votes} vs {self.all_votes}."
                  f" This comes from miscounting few crossed out candidates.")
            print()
