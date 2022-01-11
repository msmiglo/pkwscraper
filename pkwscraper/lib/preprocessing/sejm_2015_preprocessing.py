
'''import re
from zipfile import ZipFile

from lxml import html
from openpyxl import load_workbook
import xlrd'''

import json

from pkwscraper.lib.dbdriver import DbDriver
from pkwscraper.lib.preprocessing.base_preprocessing import BasePreprocessing
from pkwscraper.lib.utilities import get_parent_code


ELECTION_TYPE = "sejm"
YEAR = 2015
RAW_DATA_DIRECTORY = "./pkwscraper/data/raw/sejm/2015/"
RESCRIBED_DATA_DIRECTORY = "./pkwscraper/data/rescribed/sejm/2015/"
PREPROCESSED_DATA_DIRECTORY = "./pkwscraper/data/preprocessed/sejm/2015/"


class Region:
    # TODO - PROBABLY MOVE TO UTILITIES.
    def __init__(self, data):
        pass

    @classmethod
    def from_svg_d(cls, geo):
        data = geo
        return cls(data)

    @classmethod
    def from_json(cls, text):
        """
        Load from JSON.
        - text: str/bytes - raw content of json
        """
        data = text
        return cls(data)

    def json(self):
        """ Serialize to JSON. """
        pass

    @property
    def get_filling_boundary(self):
        """
        Get single line that defines region area with possible holes
        and separate shapes. This contains segments that join shapes
        and holes, so it is NOT suitable to draw contour of region.
        """
        pass

    @property
    def contour_lines(self):
        """
        Get set of lines that defines edge of region with possible holes
        and separate shapes. This does NOT contain segments that join
        separate shapes, so it IS suitable to draw contour of region.
        """
        pass

    """
    EXPLANATION OF GEO/REGION DATA:

    geo - this is the description of region as accepted by "d" attribute
        in "svg" HTML tag
    region - set of lines that are boundaries to region
    line - represents a closed curve, a list of points, last point is not
        needed to be equal to the first point
    part - is a part of line str, which consists of many numbers, but which
        are not separated by spaces - it is possible when numbers are negative,
        then the minus is indicating the start of new number
    point - a tuple of 2 numbers representing x and y coordinate
    number - single coordinate, especially when manipulation the geo str
        where numbers are not divided into pairs for single points
    """

    @staticmethod
    def _line_to_point(line):
        start = line.start
        return (start.real, start.imag)

    @staticmethod
    def geo_to_region(map_str):
        path = parse_path(map_str)
        curves = []
        curve = []
        for elem in path:
            if isinstance(elem, svg.path.path.Move):
                curves.append(curve)
                curve = []
            else:
                assert isinstance(elem, svg.path.path.Line)
                curve.append(elem)
        curves.append(curve)
        curves = curves[1:]
        curves = [[_line_to_point(line) for line in curve] for curve in curves]
        return curves

    @staticmethod
    def geo_to_region(geo):
        """
        dataset = [regions]
        region = [curves]
        curve = [points]
        point = [x, y]
        """
        pass


class Sejm2015Preprocessing(BasePreprocessing):
    def __init__(self):
        print("opening DB...")
        # open source DB read only
        self.source_db = DbDriver(RESCRIBED_DATA_DIRECTORY, read_only=True)
        # open target DB
        self.target_db = DbDriver(PREPROCESSED_DATA_DIRECTORY)
        print("DB opened.")

    def run_all(self):
        self._preprocess_voivodships()
        self._preprocess_okregi()
        self._preprocess_powiaty()
        self._preprocess_gminy()
        self._scale_shapes()
        self._preprocess_obwody()
        self._preprocess_protocoles()
        self._preprocess_lists()
        self._preprocess_candidates()
        self._preprocess_votes()
        self._preprocess_mandates()

        print("dumping DB tables...")
        self.target_db.dump_tables()
        print("DB closed.")
        print()

    @staticmethod
    def clean_text(text):
        """
        Used to clean non-universal texts values like names of candidates,
        committees and commissions addresses. Remove surplus spaces, replace
        non-standard quotarion marks, etc.
        """
        # TODO - MAYBE MOVE TO UTILITIES
        if text is None:
            return None
        text = " ".join(text.split())
        text = text.replace('„', '"').replace('”', '"')
        text = text.replace(' , ', ', ')
        text = text.replace('( ', '(')
        return text

    @staticmethod
    def parse_full_name(full_name):
        """
        Split full name of candidate into names and surname, asserting
        names are styled `title case` and surname is styled `all upper
        case`. Parts are separated by spaces.
        """
        # TODO - ADD SPLITTING FIRST NAME MAYBE
        # TODO - MAYBE MOVE TO UTILITIES
        names, surname = full_name.rsplit(maxsplit=1)
        i = 0
        while names.title() != names and i < 10:
            names, rest = names.rsplit(maxsplit=1)
            surname = rest + " " + surname
            i += 1
        return names, surname

    @staticmethod
    def parse_commission_address(full_address, commission_name,
                                 obwod_identifier):
        """
        Split full address of commission. It should be in format:
        full_address = `{commission_name}, {commission_address}`
        """
        clean_text = Sejm2015Preprocessing.clean_text

        full_address = clean_text(full_address)
        commission_name = clean_text(commission_name)

        # address is missing or included in commission name
        if full_address == commission_name:
            if "," in full_address:
                # name contains address
                name, address = full_address.split(",", maxsplit=1)
                name = clean_text(name)
                address = clean_text(address)
                return name, address
            else:
                # address is missing
                return commission_name, ""

        if full_address.startswith(commission_name):
            # regular case
            name = commission_name
            address = full_address[len(name):]

            if address.startswith(", "):
                # correct format
                address = address[2:]
                address = clean_text(address)
                return name, address

            if address.count(",") == 1:
                # missing one comma
                address = clean_text(address)
                return name, address

        # there is a problem with commission name
        if full_address.count(",") == 2:
            # can be determined by commas
            name, address = full_address.split(",", maxsplit=1)
            name = clean_text(name)
            address = clean_text(address)
        elif "ul. " in full_address:
            # some comma is missing, try finding the street prefix
            name, address = full_address.split("ul. ", maxsplit=1)
            name = clean_text(name)
            address = clean_text("ul. " + address)
        else:
            raise ValueError(
                f"Cannot parse adress from:\n`{full_address}`\n"
                f"`{commission_name}`\n`{address}`\n"
                f"Unit identifier:\n{obwod_identifier}\n")

        if "(" in address:
            # there is a room/hall name in the address
            address, modifier = address.split("(")
            address = clean_text(address)
            modifier = clean_text(modifier)
            name += " (" + modifier

        # # print information about malformed commission name
        # if not name.startswith(commission_name):
        #     commune_code, polling_district_number = obwod_identifier
        #     print(f"Probable typo in commision name:\n"
        #           f"{name}\n"
        #           f"{commission_name}\n"
        #           f"commune code {commune_code}, polling district no. "
        #           f"{polling_district_number}\n")

        return name, address

    @staticmethod
    def urban_or_rural(commune_full_name, commune_code):
        district_code = get_parent_code(commune_code)
        if district_code == 146500:
            return "urban"
        if commune_full_name.startswith("gm. "):
            return "rural"
        if commune_full_name.startswith("m. "):
            return "urban"
        if commune_full_name.startswith("Statki "):
            return "marine"
        if commune_full_name.startswith("Zagranica"):
            return "abroad"
        raise ValueError(f"Cannot determine `urban_or_rural` "
                         f"value from name: {commune_full_name}.")

    @staticmethod
    def merge_commune_names(partial_name, full_name, code):
        # remove urban or rural prefix
        if full_name.startswith("gm. "):
            prefix = "gm. "
            commune_name = full_name[len(prefix):]
        elif full_name.startswith("m. "):
            prefix = "m. "
            commune_name = full_name[len(prefix):]
        else:
            prefix = ""
            commune_name = full_name

        # return original full name if everything is ok
        if commune_name == partial_name:
            return full_name

        # check if full name is abbreviated
        if commune_name.endswith("."):
            commune_name = commune_name.rstrip(".")
            if partial_name.startswith(commune_name):
                # return full form if name is abbreviated
                return prefix + partial_name

        # check if old name have some suffix (the case of Słupia
        # commune, code: 260207)
        if commune_name.startswith(partial_name):
            return prefix + partial_name

        raise ValueError(
            f"Problem with names: {full_name} / {partial_name}, code={code}.")

    def _preprocess_voivodships(self):
        voivodships = self.source_db["województwa"].find({})
        self.target_db.create_table("województwa")

        for v in voivodships.values():
            code = v["code"] * 10000
            name = v["name"]
            geo = v["geo"]

            self.target_db["województwa"].put({
                "code": code,
                "name": name,
                "geo": geo,
            })

    def _preprocess_okregi(self):
        constituencies = self.source_db["okręgi"].find({})
        self.target_db.create_table("okręgi")

        for o in constituencies.values():
            number = o["number"]
            headquarters = o["headquarters"]
            voivodship_name = o["voivodship"]
            mandates = o["mandates"]
            geo = o["geo"]

            voivod_id, voivod = self.target_db["województwa"].find_one(
                {"name": voivodship_name})

            self.target_db["okręgi"].put({
                "number": number,
                "headquarters": headquarters,
                "voivodship": voivod_id,
                "mandates": mandates,
                "geo": geo,
            })

    def _preprocess_powiaty(self):
        districts = self.source_db["powiaty"].find({})
        self.target_db.create_table("powiaty")

        consituencies_numbers = self.target_db["okręgi"].find(
            query={}, fields="number")
        districts_dict = {num: list() for num in consituencies_numbers}

        for d in districts.values():
            constituency_number = d["constituency_number"]
            code = d["code"] * 100
            name = d["name"]
            geo = d.get("geo", "")

            #if code == 146500:
            #    code = 146501
            voivod_code = get_parent_code(code)
            voivod_id, voivod = self.target_db["województwa"].find_one(
                {"code": voivod_code})

            district_id = self.target_db["powiaty"].put({
                "code": code,
                "name": name,
                "geo": geo,
                "parent": voivod_id,
            })

            districts_dict[constituency_number].append(district_id)

        # assign districts to consituencies
        for constituency_number, districts_list in districts_dict.items():
            # serialize list of districts
            jsoned_districts_list = json.dumps(districts_list)

            # add field in constituency record
            con_id, con_record = self.target_db["okręgi"].find_one(
                {"number": constituency_number})
            con_record["powiat_list"] = jsoned_districts_list

            # update the record
            self.target_db["okręgi"].put(con_record, _id=con_id)

    def _preprocess_gminy(self):
        communes = self.source_db["gminy"].find({})
        self.target_db.create_table("gminy")

        code_to_name_dict = {
            code: name for code, name in self.source_db["obwody"].find(
                {}, fields=["commune_code", "commune_name"])
        }

        for c in communes.values():
            code = c["code"]
            partial_name = c["partial_name"]
            geo = c.get("geo", "")

            district_code = get_parent_code(code)
            district_id, district = self.target_db["powiaty"].find_one(
                {"code": district_code})

            full_name = code_to_name_dict[code]
            if full_name is None:
                raise ValueError(
                    f"Cannot find commune '{partial_name}' with code {code}.")
            urban_or_rural = self.urban_or_rural(full_name, code)

            merged_name = self.merge_commune_names(
                partial_name, full_name, code)

            self.target_db["gminy"].put({
                "code": code,
                "name": merged_name,
                "urban_or_rural": urban_or_rural,
                "geo": geo,
                "parent": district_id,
            })

    def _scale_shapes(self):
        pass

    def _preprocess_obwody(self):
        self.target_db.create_table("obwody")

        obwody = self.source_db["obwody"].find({})
        obwody_completion = self.source_db["obwody_uzupełnienie"].find({})

        code_to_obwod_dict = {
            (record["commune_code"], record["polling_district_number"]): _id
            for _id, record in obwody_completion.items()
        }

        gmina_code_to_id_dict = {
            code: _id
            for _id, code in self.target_db["gminy"].find(
                query={}, fields=["_id", "code"])
        }

        for ob in obwody.values():
            # get basic data
            constituency_number = ob["constituency_number"]
            senate_constituency_number = ob["senate_constituency_number"]
            commune_code = ob["commune_code"]
            commune_name = ob["commune_name"]
            polling_district_number = ob["polling_district_number"]
            full_address = ob["full_address"]
            voters = ob["voters"]

            # get completion data
            obwod_identifier = (commune_code, polling_district_number)
            ob_compl_id = code_to_obwod_dict[obwod_identifier]
            ob_compl = obwody_completion[ob_compl_id]

            completion_commune_name = ob_compl["commune_name"]
            commission_name = ob_compl.get("commission_name", "")

            # check name of commune
            if completion_commune_name != commune_name:
                raise ValueError(
                    f"Problem with commune name: {commune_name}"
                    f" / {completion_commune_name}, code: {commune_code}")

            # check name and address of commission
            try:
                name, address = self.parse_commission_address(
                    full_address, commission_name, obwod_identifier)
            except ValueError as e:
                print(e)
                print(f"---Problem with commission name and address:\n"
                      f"{full_address}\n"
                      f"{commission_name}\n"
                      f"{commune_code}\n"
                      f"{polling_district_number}\n")
                name, address = commission_name, full_address

            # preprocess additional data
            commune_id = gmina_code_to_id_dict[commune_code]
            urban_or_rural = self.urban_or_rural(commune_name, commune_code)

            # add record
            self.target_db["obwody"].put({
                "constituency": constituency_number,
                "gmina": commune_id,
                "number": polling_district_number,
                "commission_name": name,
                "address": address,
                "senate_constituency_number": senate_constituency_number,
                "urban_or_rural": urban_or_rural,
                "voters": voters,
            })

    def _preprocess_protocoles(self):
        self.target_db.create_table("protokoły")
        obwody_data = self.source_db["obwody"].find({})

        for o in obwody_data.values():
            # get identifier of polling district
            commune_code = o["commune_code"]
            polling_district_number = o["polling_district_number"]

            # TODO - MAKE DICT IN MAIN METHOD SCOPE
            commune_id = self.target_db["gminy"].find_one(
                query={"code": commune_code}, fields="_id")
            obwod_id = self.target_db["obwody"].find_one(
                query={
                    "gmina": commune_id,
                    "number": polling_district_number
                },
                fields="_id"
            )

            # get data
            voters = o["voters"]
            got_ballots = o["got_ballots"]
            unused_ballots = o["unused_ballots"]
            given_ballots = o["given_ballots"]
            proxy_voters = o["proxy_voters"]
            certificate_voters = o["certificate_voters"]
            voting_packets = o["voting_packets"]
            return_envelopes = o["return_envelopes"]
            envelopes_without_statement = o["envelopes_without_statement"]
            unsigned_statement = o["unsigned_statement"]
            without_voting_envelope = o["without_voting_envelope"]
            unseeled_voting_envelopes = o["unseeled_voting_envelopes"]
            envelopes_accepted = o["envelopes_accepted"]
            ballots_from_box = o["ballots_from_box"]
            envelopes_from_ballot_box = o["envelopes_from_ballot_box"]
            ballots_invalid = o["ballots_invalid"]
            ballots_valid = o["ballots_valid"]
            votes_invalid = o["votes_invalid"]
            invalid_2_candidates = o["invalid_2_candidates"]
            invalid_no_vote = o["invalid_no_vote"]
            invalid_candidate = o["invalid_candidate"]
            votes_valid = o["votes_valid"]

            # check correctness
            assert ballots_from_box == ballots_invalid + ballots_valid, \
                ("ballots taken", commune_code, polling_district_number)
            assert votes_valid + votes_invalid + ballots_invalid == ballots_from_box, \
                ("votes", commune_code, polling_district_number)
            assert votes_invalid >= invalid_2_candidates + invalid_no_vote + invalid_candidate, \
                ("invalid votes", commune_code, polling_district_number)
            if got_ballots != unused_ballots + given_ballots:
                print("ballots", commune_code, polling_district_number)
            if return_envelopes != envelopes_without_statement + unsigned_statement \
                + without_voting_envelope + unseeled_voting_envelopes + envelopes_accepted:
                print("envelopes", commune_code, polling_district_number)

            # add new record
            self.target_db["protokoły"].put({
                "obwod": obwod_id,
                "voters": voters,
                "got_ballots": got_ballots,
                "unused_ballots": unused_ballots,
                "given_ballots": given_ballots,
                "proxy_voters": proxy_voters,
                "certificate_voters": certificate_voters,
                "voting_packets": voting_packets,
                "return_envelopes": return_envelopes,
                "envelopes_without_statement": envelopes_without_statement,
                "unsigned_statement": unsigned_statement,
                "without_voting_envelope": without_voting_envelope,
                "unseeled_voting_envelopes": unseeled_voting_envelopes,
                "envelopes_accepted": envelopes_accepted,
                "ballots_from_box": ballots_from_box,
                "envelopes_from_ballot_box": envelopes_from_ballot_box,
                "ballots_invalid": ballots_invalid,
                "ballots_valid": ballots_valid,
                "votes_invalid": votes_invalid,
                "invalid_2_candidates": invalid_2_candidates,
                "invalid_no_vote": invalid_no_vote,
                "invalid_candidate": invalid_candidate,
                "votes_valid": votes_valid,
            })

    def _preprocess_lists(self):
        self.target_db.create_table("listy")

        lists_data = self.source_db["kandydaci"].find(
            query={}, fields=["list_number", "committee_name"])
        lists_dict = {
            self.clean_text(committee_name): list_number
            for list_number, committee_name in lists_data
        }

        committees = self.source_db["komitety"].find({})

        for c in committees.values():
            committee_number = c["number"]
            committee_symbol = c["signature"]
            committee_type = c["type"]
            committee_name = c["name"]
            committee_shortname = c.get("shortname", "")
            sejm_candidates = c["sejm_candidates"]
            senat_candidates = c["senat_candidates"]
            committee_status = c["status"]

            committee_name = self.clean_text(committee_name)
            committee_shortname = self.clean_text(committee_shortname)

            if committee_name not in lists_dict:
                continue

            list_number = lists_dict[committee_name]

            self.target_db["listy"].put({
                "committee_number": committee_number,
                "committee_type": committee_type,
                "committee_name": committee_name,
                "committee_shortname": committee_shortname,
                "committee_symbol": committee_symbol,
                "committee_status": committee_status,
                "list_number": list_number,
            })

    def _preprocess_candidates(self):
        candidates = self.source_db["kandydaci"].find({})
        self.target_db.create_table("kandydaci")

        for c in candidates.values():
            # get values
            constituency_number = c["okreg_number"]
            list_number = c["list_number"]
            committee_name = c["committee_name"]
            position = c["position"]
            surname = c["surname"]
            names = c["names"]
            gender = c["gender"]
            residence = c["residence"]
            occupation = c["occupation"]
            party = c["party"]

            # clean data
            surname = self.clean_text(surname)
            names = self.clean_text(names)
            first_name = names.split(maxsplit=1)[0]
            first_name = self.clean_text(first_name)
            committee_name = self.clean_text(committee_name)
            residence = self.clean_text(residence)
            occupation = self.clean_text(occupation)
            party = self.clean_text(party)

            # get constituency id
            constituency_id = self.target_db["okręgi"].find_one(
                query={"number": constituency_number}, fields="_id")

            # get list id
            list_id, other_list_number = self.target_db["listy"].find_one(
                query={"committee_name": committee_name},
                fields=["_id", "list_number"]
            )

            # check data correctness
            if not other_list_number == list_number:
                print(f"constituency: `{constituency_number}`")
                raise ValueError(
                    f"non-matching list number for committee"
                    f" `{committee_name}`:\n"
                    f"\t{list_number} / {other_list_number}\n"
                    f"constituency: {constituency_number}\n"
                )
            if gender not in "KM":
                raise ValueError(f"Wrong gender symbol: `{gender}`")

            # add record
            self.target_db["kandydaci"].put({
                "constituency": constituency_id,
                "list": list_id,
                "position": position,
                "surname": surname,
                "names": names,
                "first_name": first_name,
                "gender": gender,
                "residence": residence,
                "occupation": occupation,
                "party": party,
            })

    def _preprocess_votes(self):
        return
        raise NotImplementedError("Tych jeścio ni mo!")
        self.target_db.create_table("wyniki")

    def _preprocess_mandates(self):
        self.target_db.create_table("mandaty")

        mandates = self.source_db["mandaty"].find({})

        # iterate records
        for m in mandates.values():
            # get data
            constituency_number = m["constituency_number"]
            list_number = m["list_number"]
            position = m["position"]
            committee_name = m["committee_name"]
            candidate_name = m["full_name"]

            # clean texts
            committee_name = self.clean_text(committee_name)
            candidate_name = self.clean_text(candidate_name)
            first_name, surname = candidate_name.split(maxsplit=1)

            # get relations
            constituency_id = self.target_db["okręgi"].find_one(
                query={"number": constituency_number}, fields="_id")
            list_id = self.target_db["listy"].find_one(
                query={
                    "list_number": list_number,
                    "committee_name": committee_name
                },
                fields="_id"
            )

            # find candidate
            candidate_id, candidate_record = self.target_db["kandydaci"] \
                .find_one({
                    "constituency": constituency_id,
                    "list": list_id,
                    "position": position,
                })

            # check correctness
            assert candidate_record["names"].startswith(first_name)
            assert candidate_record["surname"] == surname

            # add record
            self.target_db["mandaty"].put({"candidate": candidate_id})

        # check number of mandates
        assert len(self.target_db["mandaty"].find({})) == 460


    '''
    def _download_voting_results(self):
        # votes from 41 xlsx files
        print()
        self.db.create_table("obwody_uzupełnienie")
        self.db.create_table("wyniki")

        constituencies = self.db["okręgi"].find({}, fields="number")
        for constituency_number in constituencies:
            constituency_number = int(constituency_number)
            relative_url = f"/wyniki_okr_sejm/{constituency_number:02d}.xlsx"
            self.dl.download(relative_url)
            filename = relative_url.replace("/", "_").strip("_")

            book = load_workbook(RAW_DATA_DIRECTORY + "/" + filename)
            sheet = book.active

            # read the first row with lists and candidates names
            last_protocole_column_value = sheet.cell(1, 27).value
            assert last_protocole_column_value == "Sejm - Liczba głosów ważnych oddanych łącznie na wszystkie listy kandydatów", \
                f"wrong columns alignment: {last_protocole_column_value}"
            current_list = None
            position = 0
            candidates = []
            # iterate over cells
            for column_index in range(27, sheet.max_column):
                value = sheet.cell(1, column_index+1).value
                #value = value.replace('„', '"').replace('”', '"')
                value = self.clean_text(value)
                # beginning of new list
                if not current_list:
                    current_list = value
                    candidates.append((current_list, position, "sum"))
                    continue
                # end of current list
                if value == f"Razem {current_list}":
                    position = 0
                    candidates.append((current_list, position, "sum"))
                    current_list = None
                    continue
                # get candidate
                position += 1
                candidate_name = value
                candidates.append((current_list, position, candidate_name))

            # replace candidates with their id
            def get_candidate_id(candidate):
                committee_shortname, position, candidate_name = candidate
                if position == 0:
                    return None
                committee_name = self.db["komitety"].find_one(
                    {"shortname": committee_shortname},
                    fields="name"
                )
                if committee_name is None:
                    committee_name = committee_shortname
                names, surname = self.parse_full_name(candidate_name)
                query = {
                    "okreg_number": constituency_number,
                    "committee_name": committee_name,
                    "surname": surname,
                    "names": names
                }
                results = self.db["kandydaci"].find(query, fields="_id")
                if len(results) != 1:
                    print(committee_shortname)
                    print(query)
                    print(self.db["kandydaci"].find({
                        "surname": surname, "names": names
                    }))
                    raise ValueError(f"Candidate {candidate_name} not found.")
                _id = results[0]
                return _id

            candidates = list(map(get_candidate_id, candidates))

            # check correctness
            if current_list is not None:
                raise ValueError("Did not found end of list of candidates.")

            # iterate over polling districts (1 row - 1 polling district)
            print(f"Iterating over polling districts in constituency no. {constituency_number}...")
            for row_index in range(1, sheet.max_row):
                if row_index % 30 == 0:
                    print(f"{row_index} of {sheet.max_row-1}",
                          end=", ", flush=True)
                    break  # TODO
                # get additional polling district data
                commune_name = sheet.cell(row_index+1, 2).value
                commune_code = sheet.cell(row_index+1, 3).value
                commission_name = sheet.cell(row_index+1, 4).value
                polling_district_number = sheet.cell(row_index+1, 5).value
                # put it into table
                self.db["obwody_uzupełnienie"].put({
                    "commune_name": commune_name,
                    "commune_code": commune_code,
                    "commission_name": commission_name,
                    "polling_district_number": polling_district_number
                })

                # prepare list of cells corresponding with candidates
                candidates_cell_range = list(range(27, sheet.max_column))
                if len(candidates_cell_range) != len(candidates):
                    raise ValueError(
                        f"Different lenght of candidates lists: "
                        f"{len(candidates)}/{len(candidates_cell_range)}"
                    )

                # iterate over candidates in the given polling district
                for candidate_id, column_index in zip(
                        candidates, candidates_cell_range):
                    #list_name, position, candidate_name = candidate
                    votes = sheet.cell(row_index+1, column_index+1).value
                    # put data in table
                    if candidate_id:
                        self.db["wyniki"].put({
                            "commune_code": commune_code,
                            "polling_district_number": polling_district_number,
                            "candidate": candidate_id,
                            "votes": votes
                        })
    '''




